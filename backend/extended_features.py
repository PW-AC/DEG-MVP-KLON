"""
Extended Features Module for Insurance Broker System
Includes: Provisions, Email Integration, Consultation Protocols, 
Appointments, Comparison Calculator, Export Functions
"""

from fastapi import APIRouter, HTTPException, Form, Query
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, date, timedelta
from enum import Enum
import uuid
import csv
import io
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import base64

extended_router = APIRouter(prefix="/api")

# ==================== PROVISIONSABRECHNUNG ====================

class ProvisionStatus(str, Enum):
    OFFEN = "offen"
    GEBUCHT = "gebucht"
    AUSGEZAHLT = "ausgezahlt"
    STORNIERT = "storniert"


class Provisionsabrechnung(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vertrag_id: str
    vu_id: str
    abrechnungsmonat: str  # Format: YYYY-MM
    
    # Provisionsdetails
    abschlussprovision: Optional[float] = 0
    bestandsprovision: Optional[float] = 0
    folgeprovision: Optional[float] = 0
    bonusprovision: Optional[float] = 0
    gesamtprovision: float = 0
    
    # Stornohaftung
    stornohaftungszeit_monate: Optional[int] = 24
    stornohaftung_bis: Optional[date] = None
    storno_risiko: Optional[float] = 0
    
    # Status
    status: ProvisionStatus = ProvisionStatus.OFFEN
    gebucht_am: Optional[datetime] = None
    ausgezahlt_am: Optional[datetime] = None
    
    # Referenzen
    abrechnungsnummer: Optional[str] = None
    vu_abrechnungsnummer: Optional[str] = None
    
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)


# ==================== EMAIL INTEGRATION ====================

class EmailTemplate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    subject: str
    body: str
    variables: List[str] = []  # e.g., ["kunde_name", "vertragsnummer"]
    category: str  # "reminder", "info", "contract", etc.
    created_at: datetime = Field(default_factory=datetime.utcnow)


class EmailQueue(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    recipient: str
    subject: str
    body: str
    template_id: Optional[str] = None
    status: str = "pending"  # pending, sent, failed
    scheduled_for: Optional[datetime] = None
    sent_at: Optional[datetime] = None
    error: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)


# ==================== BERATUNGSPROTOKOLLE ====================

class BeratungsprotokollTyp(str, Enum):
    ERSTBERATUNG = "erstberatung"
    FOLGEBERATUNG = "folgeberatung"
    SCHADENBERATUNG = "schadenberatung"
    VERTRAGSANPASSUNG = "vertragsanpassung"
    KUENDIGUNG = "kuendigung"


class Beratungsprotokoll(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    kunde_id: str
    berater: str
    datum: datetime = Field(default_factory=datetime.utcnow)
    typ: BeratungsprotokollTyp
    
    # Beratungsinhalt
    anlass: str
    besprochene_themen: List[str] = []
    empfehlungen: List[str] = []
    kundenwuensche: Optional[str] = None
    
    # Dokumentation nach IDD
    bedarfsanalyse_durchgefuehrt: bool = True
    geeignetheitspruefung: bool = True
    risikohinweise_gegeben: bool = True
    produktinformationen_ausgehaendigt: bool = True
    
    # Ergebnis
    angebote_erstellt: List[str] = []  # Vertrag IDs
    abgeschlossene_vertraege: List[str] = []
    naechster_termin: Optional[date] = None
    
    # Unterschriften
    kunde_unterschrift: Optional[str] = None  # Base64 signature
    berater_unterschrift: Optional[str] = None
    
    created_at: datetime = Field(default_factory=datetime.utcnow)


# ==================== TERMINVERWALTUNG ====================

class TerminTyp(str, Enum):
    BERATUNG = "beratung"
    SCHADENBESICHTIGUNG = "schadenbesichtigung"
    VERTRAGSABSCHLUSS = "vertragsabschluss"
    WIEDERVORLAGE = "wiedervorlage"
    TELEFONAT = "telefonat"
    SONSTIGES = "sonstiges"


class Termin(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    titel: str
    beschreibung: Optional[str] = None
    typ: TerminTyp
    
    # Zeitangaben
    start: datetime
    ende: datetime
    ganztaegig: bool = False
    
    # Teilnehmer
    kunde_id: Optional[str] = None
    vertrag_id: Optional[str] = None
    berater: str
    
    # Erinnerungen
    erinnerung_minuten_vorher: Optional[int] = 15
    erinnerung_gesendet: bool = False
    
    # Status
    status: str = "geplant"  # geplant, stattgefunden, abgesagt, verschoben
    
    # Wiedervorlage
    ist_wiedervorlage: bool = False
    wiedervorlage_grund: Optional[str] = None
    
    # Notizen
    notizen: Optional[str] = None
    ergebnis: Optional[str] = None
    
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)


# ==================== TARIFVERGLEICH ====================

class TarifVergleich(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    kunde_id: Optional[str] = None
    sparte: str
    
    # Vergleichsparameter
    parameter: Dict[str, Any] = {}  # Flexible parameters per insurance type
    
    # Ergebnisse
    tarife: List[Dict[str, Any]] = []  # List of tariff offers
    empfehlung: Optional[str] = None
    
    created_at: datetime = Field(default_factory=datetime.utcnow)


class TarifAngebot(BaseModel):
    vu_id: str
    vu_name: str
    tarif_name: str
    jahresbeitrag: float
    monatsbeitrag: float
    leistungen: Dict[str, Any]
    vorteile: List[str] = []
    nachteile: List[str] = []
    score: float = 0  # Comparison score 0-100


# ==================== EXPORT FUNKTIONEN ====================

@extended_router.get("/export/kunden/csv")
async def export_kunden_csv():
    """Export all customers as CSV"""
    from server import db
    
    kunden = await db.kunden.find({}).to_list(length=None)
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        'kunde_id', 'anrede', 'titel', 'vorname', 'name',
        'strasse', 'plz', 'ort', 'email', 'telefon',
        'geburtsdatum', 'created_at'
    ])
    writer.writeheader()
    
    for kunde in kunden:
        writer.writerow({
            'kunde_id': kunde.get('kunde_id'),
            'anrede': kunde.get('anrede'),
            'titel': kunde.get('titel'),
            'vorname': kunde.get('vorname'),
            'name': kunde.get('name'),
            'strasse': kunde.get('strasse'),
            'plz': kunde.get('plz'),
            'ort': kunde.get('ort'),
            'email': kunde.get('telefon', {}).get('email'),
            'telefon': kunde.get('telefon', {}).get('telefon_privat'),
            'geburtsdatum': kunde.get('persoenliche_daten', {}).get('geburtsdatum'),
            'created_at': kunde.get('created_at')
        })
    
    content = output.getvalue()
    return {
        "filename": f"kunden_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        "content": base64.b64encode(content.encode()).decode(),
        "mime_type": "text/csv"
    }


@extended_router.get("/export/vertraege/excel")
async def export_vertraege_excel():
    """Export all contracts as Excel"""
    from server import db
    
    vertraege = await db.vertraege.find({}).to_list(length=None)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Verträge"
    
    # Headers
    headers = [
        'Vertragsnummer', 'Interne Nr.', 'Kunde ID', 'Gesellschaft',
        'Produkt/Sparte', 'Tarif', 'Status', 'Beginn', 'Ablauf',
        'Beitrag Brutto', 'Beitrag Netto', 'Zahlungsweise'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, vertrag in enumerate(vertraege, 2):
        ws.cell(row=row, column=1, value=vertrag.get('vertragsnummer'))
        ws.cell(row=row, column=2, value=vertrag.get('interne_vertragsnummer'))
        ws.cell(row=row, column=3, value=vertrag.get('kunde_id'))
        ws.cell(row=row, column=4, value=vertrag.get('gesellschaft'))
        ws.cell(row=row, column=5, value=vertrag.get('produkt_sparte'))
        ws.cell(row=row, column=6, value=vertrag.get('tarif'))
        ws.cell(row=row, column=7, value=vertrag.get('vertragsstatus'))
        ws.cell(row=row, column=8, value=vertrag.get('beginn'))
        ws.cell(row=row, column=9, value=vertrag.get('ablauf'))
        ws.cell(row=row, column=10, value=vertrag.get('beitrag_brutto'))
        ws.cell(row=row, column=11, value=vertrag.get('beitrag_netto'))
        ws.cell(row=row, column=12, value=vertrag.get('zahlungsweise'))
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    content = output.getvalue()
    return {
        "filename": f"vertraege_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "content": base64.b64encode(content).decode(),
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }


@extended_router.get("/export/report/monthly")
async def generate_monthly_report(year: int = Query(default=datetime.now().year),
                                 month: int = Query(default=datetime.now().month)):
    """Generate monthly business report"""
    from server import db
    
    # Get data for the month
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    # Statistics
    total_customers = await db.kunden.count_documents({})
    new_customers = await db.kunden.count_documents({
        "created_at": {"$gte": start_date, "$lt": end_date}
    })
    
    total_contracts = await db.vertraege.count_documents({})
    new_contracts = await db.vertraege.count_documents({
        "created_at": {"$gte": start_date, "$lt": end_date}
    })
    
    # Calculate premiums
    all_contracts = await db.vertraege.find({"vertragsstatus": "aktiv"}).to_list(length=None)
    total_premium = sum(float(c.get('beitrag_brutto', 0) or 0) for c in all_contracts)
    
    # Schadensmeldungen
    total_claims = await db.schadensmeldungen.count_documents({
        "meldedatum": {"$gte": start_date, "$lt": end_date}
    })
    
    report = {
        "period": f"{year}-{month:02d}",
        "generated_at": datetime.utcnow().isoformat(),
        "summary": {
            "total_customers": total_customers,
            "new_customers_this_month": new_customers,
            "total_contracts": total_contracts,
            "new_contracts_this_month": new_contracts,
            "total_annual_premium": round(total_premium, 2),
            "claims_reported": total_claims
        },
        "charts_data": {
            "customer_growth": [],  # Would be filled with actual trend data
            "premium_distribution": [],  # Would be filled with actual distribution
            "claims_by_type": []  # Would be filled with actual claims data
        }
    }
    
    return report


# ==================== PROVISIONSABRECHNUNG ENDPOINTS ====================

@extended_router.post("/provisionen", response_model=Provisionsabrechnung)
async def create_provision(provision: Provisionsabrechnung):
    """Create new provision entry"""
    from server import db
    
    # Calculate total
    provision.gesamtprovision = (
        provision.abschlussprovision +
        provision.bestandsprovision +
        provision.folgeprovision +
        provision.bonusprovision
    )
    
    # Calculate Stornohaftung end date
    if provision.stornohaftungszeit_monate:
        vertrag = await db.vertraege.find_one({"id": provision.vertrag_id})
        if vertrag and vertrag.get('beginn'):
            beginn = datetime.fromisoformat(vertrag['beginn']) if isinstance(vertrag['beginn'], str) else vertrag['beginn']
            provision.stornohaftung_bis = beginn + timedelta(days=provision.stornohaftungszeit_monate * 30)
    
    await db.provisionen.insert_one(provision.dict())
    return provision


@extended_router.get("/provisionen/summary")
async def get_provisions_summary(year: int = Query(default=datetime.now().year)):
    """Get provisions summary for a year"""
    from server import db
    
    all_provisions = await db.provisionen.find({
        "abrechnungsmonat": {"$regex": f"^{year}"}
    }).to_list(length=None)
    
    monthly_summary = {}
    for month in range(1, 13):
        month_str = f"{year}-{month:02d}"
        month_provisions = [p for p in all_provisions if p.get('abrechnungsmonat') == month_str]
        
        monthly_summary[month_str] = {
            "abschlussprovision": sum(p.get('abschlussprovision', 0) for p in month_provisions),
            "bestandsprovision": sum(p.get('bestandsprovision', 0) for p in month_provisions),
            "folgeprovision": sum(p.get('folgeprovision', 0) for p in month_provisions),
            "bonusprovision": sum(p.get('bonusprovision', 0) for p in month_provisions),
            "gesamt": sum(p.get('gesamtprovision', 0) for p in month_provisions),
            "count": len(month_provisions)
        }
    
    return {
        "year": year,
        "monthly": monthly_summary,
        "yearly_total": sum(m['gesamt'] for m in monthly_summary.values()),
        "storno_risk": sum(p.get('storno_risiko', 0) for p in all_provisions)
    }


# ==================== EMAIL TEMPLATE ENDPOINTS ====================

@extended_router.post("/email/templates", response_model=EmailTemplate)
async def create_email_template(template: EmailTemplate):
    """Create new email template"""
    from server import db
    await db.email_templates.insert_one(template.dict())
    return template


@extended_router.get("/email/templates", response_model=List[EmailTemplate])
async def get_email_templates(category: Optional[str] = None):
    """Get all email templates"""
    from server import db
    query = {"category": category} if category else {}
    templates = await db.email_templates.find(query).to_list(length=None)
    return [EmailTemplate(**t) for t in templates]


@extended_router.post("/email/send")
async def send_email(
    recipient: str,
    subject: str,
    body: str,
    template_id: Optional[str] = None,
    variables: Optional[Dict[str, str]] = None
):
    """Send or queue an email"""
    from server import db
    
    # If template is used, replace variables
    if template_id and variables:
        template = await db.email_templates.find_one({"id": template_id})
        if template:
            body = template['body']
            subject = template['subject']
            for key, value in variables.items():
                body = body.replace(f"{{{key}}}", value)
                subject = subject.replace(f"{{{key}}}", value)
    
    email = EmailQueue(
        recipient=recipient,
        subject=subject,
        body=body,
        template_id=template_id,
        status="pending"
    )
    
    await db.email_queue.insert_one(email.dict())
    
    # Here you would integrate with actual email service
    # For now, just mark as sent
    await db.email_queue.update_one(
        {"id": email.id},
        {"$set": {"status": "sent", "sent_at": datetime.utcnow()}}
    )
    
    return {"success": True, "email_id": email.id}


# ==================== BERATUNGSPROTOKOLL ENDPOINTS ====================

@extended_router.post("/beratungsprotokolle", response_model=Beratungsprotokoll)
async def create_beratungsprotokoll(protokoll: Beratungsprotokoll):
    """Create consultation protocol"""
    from server import db
    await db.beratungsprotokolle.insert_one(protokoll.dict())
    return protokoll


@extended_router.get("/beratungsprotokolle", response_model=List[Beratungsprotokoll])
async def get_beratungsprotokolle(
    kunde_id: Optional[str] = None,
    berater: Optional[str] = None,
    typ: Optional[BeratungsprotokollTyp] = None
):
    """Get consultation protocols"""
    from server import db
    query = {}
    if kunde_id:
        query["kunde_id"] = kunde_id
    if berater:
        query["berater"] = berater
    if typ:
        query["typ"] = typ.value
    
    protokolle = await db.beratungsprotokolle.find(query).to_list(length=None)
    return [Beratungsprotokoll(**p) for p in protokolle]


# ==================== TERMIN ENDPOINTS ====================

@extended_router.post("/termine", response_model=Termin)
async def create_termin(termin: Termin):
    """Create appointment"""
    from server import db
    await db.termine.insert_one(termin.dict())
    return termin


@extended_router.get("/termine", response_model=List[Termin])
async def get_termine(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    berater: Optional[str] = None,
    kunde_id: Optional[str] = None
):
    """Get appointments"""
    from server import db
    query = {}
    
    if start_date and end_date:
        query["start"] = {
            "$gte": datetime.combine(start_date, datetime.min.time()),
            "$lte": datetime.combine(end_date, datetime.max.time())
        }
    
    if berater:
        query["berater"] = berater
    if kunde_id:
        query["kunde_id"] = kunde_id
    
    termine = await db.termine.find(query).to_list(length=None)
    return [Termin(**t) for t in termine]


@extended_router.get("/termine/wiedervorlagen")
async def get_wiedervorlagen():
    """Get all follow-ups"""
    from server import db
    wiedervorlagen = await db.termine.find({
        "ist_wiedervorlage": True,
        "status": "geplant"
    }).to_list(length=None)
    
    return [Termin(**w) for w in wiedervorlagen]


# ==================== TARIFVERGLEICH ENDPOINTS ====================

@extended_router.post("/tarifvergleich/calculate")
async def calculate_tarif_comparison(
    sparte: str,
    parameter: Dict[str, Any]
):
    """Calculate insurance tariff comparison"""
    from server import db
    
    # This would integrate with actual tariff calculation APIs
    # For demo, we create sample data
    
    sample_tarife = []
    vus = await db.vus.find({}).to_list(length=5)
    
    for vu in vus:
        base_price = 500 + (hash(vu.get('name', '')) % 1000)
        tarif = TarifAngebot(
            vu_id=vu.get('id'),
            vu_name=vu.get('name'),
            tarif_name=f"{sparte} Optimal",
            jahresbeitrag=base_price,
            monatsbeitrag=base_price / 12,
            leistungen={
                "deckungssumme": "10 Mio EUR",
                "selbstbeteiligung": "150 EUR",
                "extras": ["Auslandsschutz", "Rechtsschutz"]
            },
            vorteile=["Günstig", "Umfassender Schutz"],
            nachteile=["Hohe Selbstbeteiligung"],
            score=75 + (hash(vu.get('name', '')) % 25)
        )
        sample_tarife.append(tarif.dict())
    
    # Sort by score
    sample_tarife.sort(key=lambda x: x['score'], reverse=True)
    
    comparison = TarifVergleich(
        sparte=sparte,
        parameter=parameter,
        tarife=sample_tarife,
        empfehlung=f"Wir empfehlen {sample_tarife[0]['tarif_name']} von {sample_tarife[0]['vu_name']}"
    )
    
    await db.tarifvergleiche.insert_one(comparison.dict())
    
    return comparison


# ==================== INITIALIZE COLLECTIONS ====================

def init_extended_collections(db):
    """Initialize extended collections in the database"""
    if not hasattr(db, 'provisionen'):
        from server import SimpleCollection
        db.provisionen = SimpleCollection()
        db.email_templates = SimpleCollection()
        db.email_queue = SimpleCollection()
        db.beratungsprotokolle = SimpleCollection()
        db.termine = SimpleCollection()
        db.tarifvergleiche = SimpleCollection()
    
    # Create default email templates
    default_templates = [
        EmailTemplate(
            name="Kündigungserinnerung",
            subject="Erinnerung: Ihr Vertrag {vertragsnummer} läuft bald aus",
            body="Sehr geehrte/r {kunde_name},\n\nIhr Vertrag {vertragsnummer} läuft am {ablaufdatum} aus.\n\nBitte kontaktieren Sie uns für eine Verlängerung.\n\nMit freundlichen Grüßen",
            variables=["kunde_name", "vertragsnummer", "ablaufdatum"],
            category="reminder"
        ),
        EmailTemplate(
            name="Willkommen",
            subject="Willkommen bei uns, {kunde_name}!",
            body="Sehr geehrte/r {kunde_name},\n\nVielen Dank für Ihr Vertrauen.\n\nMit freundlichen Grüßen",
            variables=["kunde_name"],
            category="info"
        )
    ]
    
    # Add templates if not exist
    import asyncio
    async def add_templates():
        for template in default_templates:
            existing = await db.email_templates.find_one({"name": template.name})
            if not existing:
                await db.email_templates.insert_one(template.dict())
    
    asyncio.create_task(add_templates())